from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from uuid import UUID
from ...models.laboratory import LaboratorySamples


@login_required
def lab_prep_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    # Cek permission
    context = {
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-lab/preparations/list-laboratory.html',context)

class lab_prep_data(View):
    def post(self, request):
        # Ambil semua data invoice yang valid
        data_preparations = self._datatables(request)
        return JsonResponse(data_preparations, safe=False)
    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = LaboratorySamples.objects.all()

        if search:
            data = data.filter(
                Q(waybill_number__icontains=search) |
                Q(sample_id__icontains=search) 
            )
       
        # Filter berdasarkan parameter dari request
        startDate = request.POST.get('startDate')
        endDate   = request.POST.get('endDate')
        priority  = request.POST.get('priority')
        status    = request.POST.get('status')


        if startDate and endDate:
            data = data.filter(date_received__range=[startDate, endDate])

        if priority:
            data = data.filter(priority=priority)
        if status:
            data = data.filter(status=status)    

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"             : item.id,
                "date_received"  : item.date_received,
                "shift"          : item.shift,
                "priority"       : item.priority,
                "job_number"      : item.job_number,
                "waybill_number" : item.waybill_number,
                "sections"       : item.sections,
                "sample_id"      : item.sample_id,
                "mral_order"     : item.mral_order,
                "roa_order"      : item.roa_order,
                "job_number"     : item.job_number,
                "sample_wet"     : item.sample_wet,
                "sample_final"   : item.sample_final,
                "sample_press"   : item.sample_press,
                "sample_analysis": item.sample_analysis,
                "status"         : item.status,
                "remarks"        : item.remarks
                } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required
def update_prep_value(request, id):
    allowed_groups = ['superadmin', 'data-control', 'admin-mgoqa', 'admin-lab']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse({'status': 'error', 'message': 'You do not have permission'}, status=403)

    if request.method == 'POST':
        try:
            job = LaboratorySamples.objects.get(id=id)

            # Ambil nilai input (gunakan 0 jika kosong)
            sample_final    = int(request.POST.get('sample_final') or 0)
            sample_press    = int(request.POST.get('sample_press') or 0)
            sample_analysis = int(request.POST.get('sample_analysis') or 0)
            remarks         = request.POST.get('remarks')

            # Validasi berurutan
            if sample_final > 0 and (job.sample_wet or 0) <= 0:
                return JsonResponse({'error': 'Sample Wet must be > 0 to input Final'}, status=400)
            if sample_press > 0 and (job.sample_final or 0) <= 0:
                return JsonResponse({'error': 'Sample Final must be > 0 to input Press'}, status=400)
            if sample_analysis > 0 and (job.sample_press or 0) <= 0:
                return JsonResponse({'error': 'Sample Press must be > 0 to input Analysis'}, status=400)

            # Update nilai jika ada yang dimasukkan
            updated_status = job.status  # default, tidak berubah

            if sample_final > 0 and (job.sample_final or 0) == 0:
                updated_status = 'Waiting Final'
            if sample_press > 0 and (job.sample_press or 0) == 0:
                updated_status = 'Waiting Analyse'
            if sample_analysis > 0 and (job.sample_analysis or 0) == 0:
                updated_status = 'Complete'

            # Simpan nilai baru
            job.sample_final    = sample_final
            job.sample_press    = sample_press
            job.sample_analysis = sample_analysis
            job.remarks         = remarks
            job.status          = updated_status
            job.save()

            return JsonResponse({
                'id': job.id,
                'sample_wet': job.sample_wet,
                'sample_final': job.sample_final,
                'sample_press': job.sample_press,
                'sample_analysis': job.sample_analysis,
                'status': job.status,
                'remarks': job.remarks
            })

        except LaboratorySamples.DoesNotExist:
            return JsonResponse({'error': 'Data not found'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Method not allowed'}, status=405)


# @csrf_exempt
# def export_data_waybill(request):
#     start_date  = request.GET.get('startDate')
#     end_date    = request.GET.get('endDate')
#     mral_order  = request.GET.get('mral_order')  # sebelumnya pakai POST, diganti GET untuk konsistensi
#     roa_order   = request.GET.get('roa_order')

#     workbook = Workbook()
#     ws = workbook.active
#     if ws is None:
#         raise ValueError("Workbook tidak memiliki worksheet aktif")

#     worksheet: Worksheet = ws
#     worksheet.title = "Export Data Ore - Waybill"

#     # Header kolom
#     headers = [
#         'No', 'Date', 'Waybill Number', 'Qty', 'Sample Id',
#         'Mral Order', 'Roa Order', 'Remarks'
#     ]

#     # Field database yang diambil
#     fields = [
#         'delivery', 'waybill_number', 'numb_sample', 'sample_id',
#         'mral_order', 'roa_order', 'remarks'
#     ]

#     # Tulis header ke worksheet
#     for col_num, title in enumerate(headers, 1):
#         cell = worksheet.cell(row=1, column=col_num, value=title)
#         cell.font = Font(bold=True)

#     # Ambil data dari DB
#     queryset = Waybills.objects.all()
#     if start_date and end_date:
#         queryset = queryset.filter(delivery__range=[start_date, end_date])
#     if mral_order:
#         queryset = queryset.filter(mral_order=mral_order)
#     if roa_order:
#         queryset = queryset.filter(roa_order=roa_order)

#     queryset = queryset.values_list(*fields)

#     # Tulis data ke worksheet
#     for row_index, row in enumerate(queryset, start=2):
#         worksheet.cell(row=row_index, column=1, value=row_index - 1)  # No
#         for col_num, value in enumerate(row, start=2):
#             if isinstance(value, datetime):
#                 value = value.replace(tzinfo=None)
#             worksheet.cell(row=row_index, column=col_num, value=value)

#     # Sesuaikan lebar kolom
#     for col_num, title in enumerate(headers, 1):
#         col_letter = get_column_letter(col_num)
#         max_length = len(str(title))
#         for cell in worksheet[col_letter]:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         worksheet.column_dimensions[col_letter].width = max_length + 2

#     # Buat file response
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="data_waybill.xlsx"'
#     workbook.save(response)
#     return response
