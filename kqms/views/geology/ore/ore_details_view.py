from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.details_mral_view import DetailsMral
from ....models.details_roa_view import DetailsRoa
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime


@login_required
def ore_details_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    # Cek permission
    context = {
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-mgoqa/production-ore/list-ore-details.html',context)


class OreDetailsView(View):

    def post(self, request):
        # Ambil semua data invoice yang valid
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = DetailsMral.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(prospect_area__icontains=search) |
                Q(mine_block__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(ore_class__icontains=search) |
                Q(grade_control__icontains=search) |
                Q(unit_truck__icontains=search) |
                Q(stockpile__icontains=search) |
                Q(pile_id__icontains=search) |
                Q(batch_code__icontains=search) |
                Q(batch_status__icontains=search) |
                Q(truck_factor__icontains=search) |
                Q(sample_number__icontains=search) 
            )
       

        # Filter berdasarkan parameter dari request
        startDate       = request.POST.get('startDate')
        endDate         = request.POST.get('endDate')
        material_filter = request.POST.get('material_filter')
        batch_status    = request.POST.get('batch_status')
        area_filter     = request.POST.get('area_filter')
        point_filter    = request.POST.get('point_filter')
        source_filter   = request.POST.get('source_filter')

        if startDate and endDate:
            data = data.filter(tgl_production__range=[startDate, endDate])

        if material_filter:
            data = data.filter(nama_material=material_filter)

        if batch_status:
            data = data.filter(batch_status=batch_status)

        if area_filter:
            data = data.filter(stockpile=area_filter)

        if point_filter:
            data = data.filter(pile_id=point_filter)

        if source_filter:
            data = data.filter(prospect_area=source_filter)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [

            {
                "tgl_production": item.tgl_production,
                "category": item.category,
                "shift": item.shift,
                "prospect_area": item.prospect_area,
                "mine_block": item.mine_block,
                "rl": (str(item.from_rl) if item.from_rl is not None else '') + '-' + (str(item.to_rl) if item.to_rl is not None else ''),
                # "to_rl": ,
                "nama_material": item.nama_material,
                "ore_class": item.ore_class,
                "ni_grade": item.ni_grade,
                "grade_control": item.grade_control,
                "unit_truck": item.unit_truck,
                "stockpile": item.stockpile,
                "pile_id": item.pile_id,
                "batch_code": item.batch_code,
                "increment": item.increment,
                "batch_status": item.batch_status,
                "ritase": item.ritase,
                # "tonnage": item.tonnage,
                 "tonnage": round(item.tonnage, 2),
                "pile_status": item.pile_status,
                "truck_factor": item.truck_factor,
                "remarks": item.remarks,
                "sample_number": item.sample_number,
                "mral_ni": item.mral_ni
                
            } for item in object_list
        ]

        return {
            'draw': draw,
            'recordsTotal': records_total,
            'recordsFiltered': total_records_filtered,
            'data': data,
            'start': start,
            'length': length,
            'totalPages': total_pages,
        }
    
@csrf_exempt
@csrf_exempt
def export_detail_mral(request):
    # Ambil parameter filter dari request
    start_date     = request.GET.get('startDate')
    end_date       = request.GET.get('endDate')
    material       = request.GET.get('material_filter')
    batch_status   = request.GET.get('batch_status')
    area_filter    = request.GET.get('area_filter')
    point_filter   = request.GET.get('point_filter')
    source_filter  = request.GET.get('source_filter')

    workbook = Workbook()
    ws = workbook.active
    if ws is None:
        raise ValueError("Workbook tidak memiliki worksheet aktif")
    
    worksheet: Worksheet = ws
    worksheet.title = "Export Data Ore - MRAL"

    # Header kolom Excel
    headers = [
        'No', 'Date', 'Shift', 'Source', 'Block', 'From', 'To-RL', 'Material', 'Class',
        'Ni-Expect', 'Mine Geology', 'Units', 'Stockpile', 'Dome', 'Batch', 'Increment',
        'Batch Status', 'Ritase', 'Tonnage', 'Dome Status', 'Truck Factor', 'Sample ID',
        'Remarks', 'Ni', 'Co', 'Fe2O3', 'Fe', 'MgO', 'SiO2', 'Sm'
    ]

    fields = [
        'tgl_production', 'shift', 'prospect_area', 'mine_block', 'from_rl', 'to_rl',
        'nama_material', 'ore_class', 'ni_grade', 'grade_control', 'unit_truck',
        'stockpile', 'pile_id', 'batch_code', 'increment', 'batch_status', 'ritase',
        'tonnage', 'pile_status', 'truck_factor', 'sample_number', 'remarks',
        'mral_ni', 'mral_co', 'mral_fe2o3', 'mral_fe', 'mral_mgo', 'mral_sio2', 'mral_sm'
    ]

    # Filter queryset
    queryset = DetailsMral.objects.all()
    if start_date and end_date:
        queryset = queryset.filter(tgl_production__range=[start_date, end_date])
    if material:
        queryset = queryset.filter(nama_material=material)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)

    queryset = queryset.values_list(*fields)

    # Tulis header
    for col_num, title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    # Tulis data
    for row_index, row in enumerate(queryset, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)  # Kolom No
        for col_num, value in enumerate(row, start=2):
            worksheet.cell(row=row_index, column=col_num, value=value)

    # Atur lebar kolom otomatis
    for col_num, title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(title))
        for cell in worksheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Buat response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Ore_data_mral.xlsx"'
    workbook.save(response)
    return response

def export_detail_roa(request):
    # Ambil parameter filter dari request
    start_date     = request.GET.get('startDate')
    end_date       = request.GET.get('endDate')
    material       = request.GET.get('material_filter')
    batch_status   = request.GET.get('batch_status')
    area_filter    = request.GET.get('area_filter')
    point_filter   = request.GET.get('point_filter')
    source_filter  = request.GET.get('source_filter')

    workbook = Workbook()
    ws = workbook.active
    if ws is None:
        raise ValueError("Workbook tidak memiliki worksheet aktif")

    worksheet: Worksheet = ws
    worksheet.title = "Export Data Ore - ROA"

    # Header dan field
    headers = [
        'No', 'Date', 'Shift', 'Source', 'Block', 'From', 'To-Rl', 'Material',
        'Class', 'Ni-Expect', 'Mine Geology', 'Units', 'Stockpile', 'Dome', 'Batch',
        'Increment', 'Batch Status', 'Ritase', 'Tonnage', 'Dome Status', 'Sample Id',
        'Remarks', 'Ni', 'Co', 'Al2O3', 'CaO', 'Cr2O3', 'Fe2O3', 'Fe', 'MgO',
        'SiO2', 'Sm', 'Mc'
    ]

    fields = [
        'tgl_production', 'shift', 'prospect_area', 'mine_block', 'from_rl', 'to_rl',
        'nama_material', 'ore_class', 'ni_grade', 'grade_control', 'unit_truck',
        'stockpile', 'pile_id', 'batch_code', 'increment', 'batch_status', 'ritase',
        'tonnage', 'pile_status', 'sample_number', 'remarks',
        'roa_ni', 'roa_co', 'roa_al2o3', 'roa_cao', 'roa_cr2o3', 'roa_fe2o3',
        'roa_fe', 'roa_mgo', 'roa_sio2', 'roa_sm', 'roa_mc'
    ]

    # Filter data
    queryset = DetailsRoa.objects.all()
    if start_date and end_date:
        queryset = queryset.filter(tgl_production__range=[start_date, end_date])
    if material:
        queryset = queryset.filter(nama_material=material)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)

    queryset = queryset.values_list(*fields)

    # Tulis header
    for col_num, title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    # Tulis data
    for row_index, row in enumerate(queryset, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)  # Kolom No
        for col_num, value in enumerate(row, start=2):
            worksheet.cell(row=row_index, column=col_num, value=value)

    # Auto-width kolom
    for col_num, title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(title))
        for cell in worksheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Buat response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Ore_data_roa.xlsx"'
    workbook.save(response)
    return response

