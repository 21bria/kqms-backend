from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.ore_productions import OreProductions
from ....models.ore_production_view import OreProductionsView
from ....models.details_mral_view import DetailsMral
from ....models.details_roa_view import DetailsRoa
from ....models.block_model import Block
from ....models.source_model import SourceMinesLoading,SourceMinesDumping,SourceMinesDome
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Count, Sum
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from ....utils.utils import clean_string
from ....models.materials import Material
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def format_angka(jumlah):
    if jumlah >= 1_000_000_000:
        return f"{jumlah / 1_000_000_000:.2f} B"
    elif jumlah >= 1_000_000:
        return f"{jumlah / 1_000_000:.2f} M"
    elif jumlah >= 1_000:
        return f"{jumlah / 1_000:.2f} K"
    else:
        return str(jumlah)


@login_required
def ore_production_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    material = Material.objects.filter(nama_material__in=['LIM', 'SAP'])

    # Cek permission
    context = {
        'material'   : material,
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-mgoqa/production-ore/list-ore.html',context)

class OreProduction(View):

    def post(self, request):
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = OreProductionsView.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(prospect_area__icontains=search) |
                Q(mine_block__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(ore_class__icontains=search) |
                Q(grade_control__icontains=search) |
                Q(unit_truck__icontains=search) |
                Q(stockpile__icontains=search) |
                Q(pile_id__icontains=search) |
                Q(batch_code__icontains=search) |
                Q(batch_status__icontains=search) |
                Q(truck_factor__icontains=search) |
                Q(sample_number__icontains=search) |
                Q(username__icontains=search) 
            )
       

        # Filter berdasarkan parameter dari request
        startDate       = request.POST.get('startDate')
        endDate         = request.POST.get('endDate')
        material_filter = request.POST.get('material_filter')
        batch_status    = request.POST.get('batch_status')
        area_filter     = request.POST.get('area_filter')
        point_filter    = request.POST.get('point_filter')
        source_filter   = request.POST.get('source_filter')

        if startDate and endDate:
            data = data.filter(tgl_production__range=[startDate, endDate])

        if material_filter:
            data = data.filter(nama_material=material_filter)

        if batch_status:
            data = data.filter(batch_status=batch_status)

        if area_filter:
            data = data.filter(stockpile=area_filter)

        if point_filter:
            data = data.filter(pile_id=point_filter)

        if source_filter:
            data = data.filter(prospect_area=source_filter)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"            : item.id,
                "tgl_production": item.tgl_production,
                "category"      : item.category,
                "shift"         : item.shift,
                "prospect_area" : item.prospect_area,
                "mine_block"    : item.mine_block,
                "rl"            : (str(item.from_rl) if item.from_rl is not None else '') + '-' + (str(item.to_rl) if item.to_rl is not None else ''),
                "nama_material" : item.nama_material,
                "ore_class"     : item.ore_class,
                "ni_grade"      : item.ni_grade,
                "grade_control" : item.grade_control,
                "unit_truck"    : item.unit_truck,
                "stockpile"     : item.stockpile,
                "pile_id"       : item.pile_id,
                "batch_code"    : item.batch_code,
                "increment"     : item.increment,
                "batch_status"  : item.batch_status,
                "ritase"        : item.ritase,
                "tonnage"       : item.tonnage,
                "pile_status"   : item.pile_status,
                "truck_factor"  : item.truck_factor,
                "remarks"       : item.remarks,
                "sample_number" : item.sample_number,
                "username"      : item.username
                
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required()    
@csrf_exempt
def export_ore_data(request):
    # Ambil parameter filter dari request
    start_date     = request.GET.get('startDate')
    end_date       = request.GET.get('endDate')
    material       = request.GET.get('material_filter')
    batch_status   = request.GET.get('batch_status')
    area_filter    = request.GET.get('area_filter')
    point_filter   = request.GET.get('point_filter')
    source_filter  = request.GET.get('source_filter')

    workbook = Workbook()
    ws = workbook.active
    if ws is None:
        raise ValueError("Workbook tidak memiliki worksheet aktif")

    worksheet: Worksheet = ws
    worksheet.title = "Export Data Ore"

    # Header kolom Excel
    headers = [
        'No', 'Date', 'Shift', 'Source', 'Block', 'From', 'To-Rl', 'Material',
        'Class', 'Ni-Expect', 'Mine Gelology', 'Units', 'Stockpile', 'Dome',
        'Batch', 'Increment', 'Batch Status', 'Ritase', 'Tonnage', 'Dome Status',
        'Truck Factor', 'Sample Id', 'Remarks'
    ]

    fields = [
        'tgl_production', 'shift', 'prospect_area', 'mine_block', 'from_rl', 'to_rl',
        'nama_material', 'ore_class', 'ni_grade', 'grade_control', 'unit_truck',
        'stockpile', 'pile_id', 'batch_code', 'increment', 'batch_status', 'ritase',
        'tonnage', 'pile_status', 'truck_factor', 'sample_number', 'remarks',
    ]

    # Tulis header ke worksheet
    for col_num, title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    # Ambil queryset dengan filter
    queryset = OreProductionsView.objects.values_list(*fields)
    if start_date and end_date:
        queryset = queryset.filter(tgl_production__range=[start_date, end_date])
    if material:
        queryset = queryset.filter(nama_material=material)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)

    # Tulis data ke worksheet
    for row_index, row in enumerate(queryset, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)  # No
        for col_num, value in enumerate(row, start=2):
            worksheet.cell(row=row_index, column=col_num, value=value)

    # Auto lebar kolom
    for col_num, title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(title))
        for cell in worksheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Buat response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ore_data.xlsx"'
    workbook.save(response)
    return response

@login_required()
def total_ore(request):
    # queryset = OreProductions.objects.exclude(id_stockpile=66)
    queryset = OreProductionsView.objects.exclude(stockpile='Temp-Rompile_KM09')

    start_date  = request.GET.get('startDate')
    end_date    = request.GET.get('endDate')

    material_filter = request.GET.get('material_filter')
    batch_status = request.GET.get('batch_status')
    area_filter = request.GET.get('area_filter')
    point_filter = request.GET.get('point_filter')
    source_filter = request.GET.get('source_filter')

    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
        queryset   = queryset.filter(tgl_production__range=[start_date, end_date])

    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)    

    result = queryset.aggregate(
        qty     = Count('*'),
        tonnage = Sum('tonnage', default=0)
    )


    return JsonResponse({
        'Qty': result['qty'],
        'Tonnage': result['tonnage']
    })

@login_required()
def total_details_mral(request):
    queryset = DetailsMral.objects.exclude(stockpile='Temp-Rompile_KM09')

    start_date  = request.GET.get('startDate')
    end_date    = request.GET.get('endDate')

    material_filter = request.GET.get('material_filter')
    batch_status = request.GET.get('batch_status')
    area_filter = request.GET.get('area_filter')
    point_filter = request.GET.get('point_filter')
    source_filter = request.GET.get('source_filter')

    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
        queryset   = queryset.filter(tgl_production__range=[start_date, end_date])

    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter) 
        
    result = queryset.aggregate(
        qty     = Count('*'),
        tonnage = Sum('tonnage', default=0)
    )
    return JsonResponse({
        'Qty': result['qty'],
        'Tonnage': result['tonnage']
    })

@login_required()
def total_details_roa(request):
    queryset = DetailsRoa.objects.exclude(stockpile='Temp-Rompile_KM09')

    start_date  = request.GET.get('startDate')
    end_date    = request.GET.get('endDate')
    material_filter = request.GET.get('material_filter')
    batch_status = request.GET.get('batch_status')
    area_filter = request.GET.get('area_filter')
    point_filter = request.GET.get('point_filter')
    source_filter = request.GET.get('source_filter')

    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date   = datetime.strptime(end_date, '%Y-%m-%d').date()
        queryset   = queryset.filter(tgl_production__range=[start_date, end_date])

    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if batch_status:
        queryset = queryset.filter(batch_status=batch_status)
    if area_filter:
        queryset = queryset.filter(stockpile=area_filter)
    if point_filter:
        queryset = queryset.filter(pile_id=point_filter)
    if source_filter:
        queryset = queryset.filter(prospect_area=source_filter)      

    result = queryset.aggregate(
        qty     = Count('*'),
        tonnage = Sum('tonnage', default=0)
    )

    return JsonResponse({
        'Qty'    : result['qty'],
        'Tonnage': result['tonnage']
    })

@login_required
def getIdOre(request, id):
    # allowed_groups = ['superadmin','data-control','admin-mgoqa']
    # if not request.user.groups.filter(name__in=allowed_groups).exists():
    #     return JsonResponse(
    #         {'status': 'error', 'message': 'You do not have permission'}, 
    #         status=403
    # )
    if request.method == 'GET':
        try:
            items = OreProductions.objects.get(id=id)

            block_name    = None
            loading_point = None
            id_dumping    = None
            dome_point    = None

            if items.id_block:
                # Ambil nama blok berdasarkan `id_block`
                block = Block.objects.filter(id=items.id_block).first()
                if block:
                    block_name = block.mine_block

            if items.id_prospect_area:
                loading = SourceMinesLoading.objects.filter(id=items.id_prospect_area).first()
                if loading:
                    loading_point = loading.loading_point

            if items.id_pile:
                dome = SourceMinesDome.objects.filter(id=items.id_pile).first()
                if dome:
                    dome_point = dome.pile_id 
                    id_dumping = dome.id_dumping 

            data = {
                'id'              : items.id,
                'tgl_production'  : items.tgl_production, 
                'category'        : clean_string(items.category),
                'shift'           : clean_string(items.shift),
                'id_prospect_area': items.id_prospect_area,
                'loading_point'   : clean_string(loading_point),
                'id_block'        : items.id_block,
                'block_name'      : clean_string(block_name),  # Tambahkan nama blok ke response
                'from_rl'         : clean_string(items.from_rl),
                'to_rl'           : clean_string(items.to_rl),
                'id_material'     : items.id_material,
                'grade_expect'    : items.grade_expect,
                'grade_control'   : clean_string(items.grade_control),
                'unit_truck'      : clean_string(items.unit_truck),
                'id_stockpile'    : id_dumping,
                'id_pile'         : items.id_pile,
                'dome_point'      : clean_string(dome_point),
                'batch_code'      : clean_string(items.batch_code),
                'increment'       : items.increment,
                'batch_status'    : clean_string(items.batch_status),
                'ritase'          : items.ritase,
                'tonnage'         : items.tonnage,
                'pile_status'     : clean_string(items.pile_status),
                'ore_class'       : clean_string(items.ore_class),
                'truck_factor'    : clean_string(items.truck_factor),
                'no_production'   : clean_string(items.no_production),
                'sale_adjust'     : clean_string(items.sale_adjust),
                'remarks'         : clean_string(items.remarks)
            }
            return JsonResponse(data)
        except OreProductions.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def delete_ore(request):
    # allowed_groups = ['superadmin']
    # if not request.user.groups.filter(name__in=allowed_groups).exists():
    #     return JsonResponse(
    #         {'status': 'error', 'message': 'You do not have permission'}, 
    #         status=403
    # )

    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if job_id:
            # Lakukan penghapusan berdasarkan ID di sini
            data = OreProductions.objects.get(id=int(job_id))
            data.delete()
            return JsonResponse({'status': 'deleted'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
