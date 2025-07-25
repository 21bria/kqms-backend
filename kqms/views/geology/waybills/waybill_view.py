from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from uuid import UUID
from ....utils.utils import clean_string
from ....models.waybill_model import Waybills
from ....models.waybill_view import listWaybills

@login_required
def waybill_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan
    # Cek permission
    context = {
        'start_date' : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'   : today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-mgoqa/waybills/list-waybills.html',context)

class Waybill_data(View):
    def post(self, request):
        # Ambil semua data invoice yang valid
        data_waybill = self._datatables(request)
        return JsonResponse(data_waybill, safe=False)
    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = listWaybills.objects.all()

        if search:
            data = data.filter(
                Q(waybill_number__icontains=search) |
                Q(sample_id__icontains=search) 
            )
       
        # Filter berdasarkan parameter dari request
        startDate = request.POST.get('startDate')
        endDate   = request.POST.get('endDate')
        mralOrder = request.POST.get('mralOrder')
        roaOrder  = request.POST.get('roaOrder')


        if startDate and endDate:
            data = data.filter(tgl_deliver__range=[startDate, endDate])

        if mralOrder:
            data = data.filter(mral_order=mralOrder)
        if roaOrder:
            data = data.filter(roa_order=roaOrder)    

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"            : item.id,
                "tgl_deliver"   : item.tgl_deliver,
                "delivery_time" : item.delivery_time,
                "waybill_number": item.waybill_number,
                "numb_sample"   : item.numb_sample,
                "sample_id"     : item.sample_id,
                "mral_order"    : item.mral_order,
                "roa_order"     : item.roa_order,
                "remarks"       : item.remarks,
                "username"      : item.username,
                } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required
def getIdWaybill(request, id):
    allowed_groups = ['superadmin','data-control','admin-mgoqa']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    if request.method == 'GET':
        try:
            items = Waybills.objects.get(id=id)
            data = {
                'id'            : items.id,
                'tgl_deliver'   : items.tgl_deliver, 
                'delivery_time' : items.delivery_time,
                'waybill_number': clean_string(items.waybill_number),
                'numb_sample'   : items.numb_sample,
                'sample_id'     : clean_string(items.sample_id),
                'mral_order'    : clean_string(items.mral_order),
                'roa_order'     : clean_string(items.roa_order),
                'remarks'       : clean_string(items.remarks)
            }
            return JsonResponse(data)
        except Waybills.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def update_waybill(request, id):
    allowed_groups = ['superadmin', 'data-control', 'admin-mgoqa']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
        )
    if request.method == 'POST':
        try:
            job = Waybills.objects.get(id=id)

            # Mengambil data dari request
            tgl_deliver   = request.POST.get('tgl_deliver')
            delivery_time = request.POST.get('delivery_time')

            # Menggabungkan tgl_deliver dan delivery_time menjadi datetime
            if tgl_deliver and delivery_time:
                try:
                    delivery = datetime.strptime(f"{tgl_deliver} {delivery_time}", "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    return JsonResponse({'error': 'Format tgl_deliver atau delivery_time salah'}, status=400)
            else:
                delivery = None

            # Menyimpan data ke dalam database
            job.tgl_deliver = tgl_deliver
            job.delivery_time = delivery_time
            job.delivery = delivery
            job.waybill_number = request.POST.get('waybill_number')
            job.numb_sample = request.POST.get('numb_sample')
            job.sample_id = request.POST.get('sample_id')
            job.mral_order = request.POST.get('mral_order')
            job.roa_order = request.POST.get('roa_order')
            job.remarks = request.POST.get('remarks')
            
            job.save()

            return JsonResponse({
                'id': job.id,
                'tgl_deliver'   : job.tgl_deliver,
                'delivery_time' : job.delivery_time,
                'delivery'      : job.delivery,
                'waybill_number': job.waybill_number,
                'numb_sample'   : job.numb_sample,
                'sample_id'     : job.sample_id,
                'mral_order'    : job.mral_order,
                'roa_order'     : job.roa_order,
                'remarks'       : job.remarks,
                'created_at'    : job.created_at
            })

        except Waybills.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Metode tidak diizinkan'}, status=405)

@login_required
def delete_waybill(request):
    allowed_groups = ['superadmin']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
        )

    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if not job_id:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'}, status=400)

        try:
            job_uuid = UUID(job_id)  # Validasi UUID format
            data = Waybills.objects.get(id=job_uuid)
            data.delete()
            return JsonResponse({'status': 'deleted'})
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid UUID format'}, status=400)
        except Waybills.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Data not found'}, status=404)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)
    

@csrf_exempt
def export_data_waybill(request):
    start_date  = request.GET.get('startDate')
    end_date    = request.GET.get('endDate')
    mral_order  = request.GET.get('mral_order')  # sebelumnya pakai POST, diganti GET untuk konsistensi
    roa_order   = request.GET.get('roa_order')

    workbook = Workbook()
    ws = workbook.active
    if ws is None:
        raise ValueError("Workbook tidak memiliki worksheet aktif")

    worksheet: Worksheet = ws
    worksheet.title = "Export Data Ore - Waybill"

    # Header kolom
    headers = [
        'No', 'Date', 'Waybill Number', 'Qty', 'Sample Id',
        'Mral Order', 'Roa Order', 'Remarks'
    ]

    # Field database yang diambil
    fields = [
        'delivery', 'waybill_number', 'numb_sample', 'sample_id',
        'mral_order', 'roa_order', 'remarks'
    ]

    # Tulis header ke worksheet
    for col_num, title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    # Ambil data dari DB
    queryset = Waybills.objects.all()
    if start_date and end_date:
        queryset = queryset.filter(delivery__range=[start_date, end_date])
    if mral_order:
        queryset = queryset.filter(mral_order=mral_order)
    if roa_order:
        queryset = queryset.filter(roa_order=roa_order)

    queryset = queryset.values_list(*fields)

    # Tulis data ke worksheet
    for row_index, row in enumerate(queryset, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)  # No
        for col_num, value in enumerate(row, start=2):
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            worksheet.cell(row=row_index, column=col_num, value=value)

    # Sesuaikan lebar kolom
    for col_num, title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(title))
        for cell in worksheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Buat file response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="data_waybill.xlsx"'
    workbook.save(response)
    return response
