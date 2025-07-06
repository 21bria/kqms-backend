from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.samples_data_view import samplesNoOrders
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet


# Index Page
@login_required
def samples_pending_page(request):
     # Ambil permissions dinamis dari database
    return render(request, 'admin-mgoqa/production-samples/list-sample-pending.html')

class SamplesPending(View):

    def post(self, request):
        # Ambil semua data invoice yang valid
        data_sample = self._datatables(request)
        return JsonResponse(data_sample, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = samplesNoOrders.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(sampling_area__icontains=search) |
                Q(sampling_point__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(batch_code__icontains=search) |
                Q(sample_number__icontains=search) |
                Q(no_sample__icontains=search)
            )
       

        # Filter berdasarkan parameter dari request
        from_date       = request.POST.get('from_date')
        to_date         = request.POST.get('to_date')
        materialFilter  = request.POST.get('materialFilter')
        typeFilter      = request.POST.get('typeFilter')
        areaFilter      = request.POST.get('areaFilter')
        pointFilter     = request.POST.get('pointFilter')


        if from_date and to_date:
            data = data.filter(tgl_sample__range=[from_date, to_date])

        if typeFilter:
            data = data.filter(type_sample=typeFilter)

        if materialFilter:
            data = data.filter(nama_material=materialFilter)

        if areaFilter:
            data = data.filter(sampling_area=areaFilter)

        if pointFilter:
            data = data.filter(sampling_point=pointFilter)


        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "tgl_sample"     : item.tgl_sample,
                "shift"          : item.shift,
                "type_sample"    : item.type_sample,
                "sample_method"  : item.sample_method,
                "sampling_area"  : item.sampling_area,
                "sampling_point" : item.sampling_point,
                "nama_material"  : item.nama_material,
                "batch_code"     : item.batch_code,
                "increments"     : item.increments,
                "size"           : item.size,
                "sample_weight"  : item.sample_weight,
                "sample_number"  : item.sample_number,
                "remark"         : item.remark,
                "sampling_deskripsi": item.sampling_deskripsi,
                "waybill_number"   : item.waybill_number
                
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@csrf_exempt
def export_samples_pending(request):
    # Ambil filter dari request
    from_date       = request.GET.get('from_date')
    to_date         = request.GET.get('to_date')
    material_filter = request.GET.get('materialFilter')
    type_filter     = request.GET.get('typeFilter')

    # Siapkan workbook
    workbook = Workbook()
    ws = workbook.active
    if ws is None:
        raise ValueError("Workbook tidak memiliki worksheet aktif")

    worksheet: Worksheet = ws
    worksheet.title = 'Export Samples Pending'

    # Header kolom
    headers = [
        'No', 'Date', 'Week', 'Month', 'Year', 'Shift', 'Type', 'Method',
        'Material', 'Sampling Area', 'Sampling Point', 'Batch', 'Increments',
        'Sample Id', 'Sample weight', 'Primer raw', 'Duplicate raw',
        'Sampling Description', 'Remark', 'Waybill'
    ]

    fields = [
        'tgl_sample', 'minggu', 'bulan', 'tahun', 'shift', 'type_sample', 'sample_method',
        'nama_material', 'sampling_area', 'sampling_point', 'batch_code', 'increments',
        'sample_number', 'sample_weight', 'primer_raw', 'duplicate_raw',
        'sampling_deskripsi', 'remark', 'waybill_number'
    ]

    # Tulis header
    for col_num, title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    # Query data
    queryset = samplesNoOrders.objects.all()
    if from_date and to_date:
        queryset = queryset.filter(tgl_sample__range=[from_date, to_date])
    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)
    if type_filter:
        queryset = queryset.filter(type_sample=type_filter)

    queryset = queryset.values_list(*fields)

    # Tulis data ke worksheet
    for row_index, row in enumerate(queryset, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)  # No
        for col_num, value in enumerate(row, start=2):
            worksheet.cell(row=row_index, column=col_num, value=value)

    # Auto lebar kolom
    for col_num, title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(title))
        for cell in worksheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample-pending-lab.xlsx"'
    workbook.save(response)
    return response

