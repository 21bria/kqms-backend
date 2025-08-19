from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ....models.selling_barging_temp import SellingBargingTemp
from ....models.selling_details_barging_view import SellingDetailsBargingTempView
from ....models.materials import Material
from ....models.master_barge import BargeUnits
from ....models.source_model import SourceMines,SourceMinesLoading,SourceMinesDumping,SourceMinesDome
from ....models.selling_code import SellingCode
from django.db.models import Sum, Count
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.core.exceptions import ValidationError
from django.core.exceptions import ObjectDoesNotExist
from django.views import View
from datetime import datetime
from django.db.models import Case, When, Value, IntegerField
from uuid import UUID

@login_required
def sale_entry_page(request):
    today = datetime.today()
    context = {
        'today': today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-selling/data/entry-list-selling.html',context)

def form_entry_page(request):
    today = datetime.today()
    context = {
        'today': today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-selling/data/form-entry.html',context)

class SellingTemp(View):
    def post(self, request):
        # Ambil semua data invoice yang valid
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = SellingDetailsBargingTempView.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(dome__icontains=search) |
                Q(unit_code__icontains=search) |
                Q(material__icontains=search) |
                Q(code_lot__icontains=search) 
            )
       

        # Filter berdasarkan parameter dari request
        from_date       = request.POST.get('startDate')
        to_date         = request.POST.get('endDate')
        materialFilter  = request.POST.get('materialFilter')
        productFilter   = request.POST.get('productFilter')

        if from_date and to_date:
            data = data.filter(date_hauling__range=[from_date, to_date])

        if materialFilter:
            data = data.filter(material=materialFilter)

        if productFilter:
            data = data.filter(code_lot=productFilter)

        columns_map = {
            1: 'date_hauling',
            2: 'time_hauling',
            3: 'barge_code',
            4: 'shift',
            5: 'dome',
            6: 'stockpile',
            7: 'material',
            8: 'unit_code',
            9: 'tonnage',
            10: 'code_lot',
            11: 'code_inc',
            12: 'code_sub',
            13: 'type_selling',
            14: 'no_urut',
        }


        order_column = int(datatables.get('order[0][column]', -1))
        order_dir = datatables.get('order[0][dir]', 'asc')

        if order_column in columns_map:
            field_name = columns_map[order_column]
            order_by = f'-{field_name}' if order_dir == 'desc' else field_name
            data = data.order_by(order_by)
        else:
            data = data.annotate(
                shift_order=Case(
                    When(shift="Day", then=Value(1)),
                    When(shift="Night", then=Value(2)),
                    default=Value(3),
                    output_field=IntegerField(),
                )
            ).order_by('-date_hauling', 'shift_order', 'time_hauling', 'no_urut')

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"           : item.id,
                "date_hauling" : item.date_hauling,
                "time_hauling" : item.time_hauling,
                "barge_code"   : item.barge_code,
                "shift"        : item.shift,
                "dome"         : item.dome,
                "stockpile"    : item.stockpile,
                "material"     : item.material,
                "unit_code"    : item.unit_code,
                "tonnage"      : item.tonnage,
                "code_lot"     : item.code_lot,
                "code_inc"     : item.code_inc,
                "code_sub"     : item.code_sub,
                "type_selling" : item.type_selling,
                "no_urut"      : item.no_urut
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required
def total_selling(request):
    data = SellingDetailsBargingTempView.objects.all()

    # Ambil parameter dari request
    from_date       = request.GET.get('startDate')
    to_date         = request.GET.get('endDate')
    materialFilter  = request.GET.get('materialFilter')
    areaFilter      = request.GET.get('areaFilter')
    pointFilter     = request.GET.get('pointFilter')
    factoriesFilter = request.GET.get('factoriesFilter')
    productFilter   = request.GET.get('productFilter')

    # Filter berdasarkan tanggal
    if from_date and to_date:
        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            data = data.filter(date_hauling__range=[from_date, to_date])
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

    # Filter berdasarkan parameter lain
    if materialFilter:
        data = data.filter(material=materialFilter)
    if areaFilter:
        data = data.filter(stockpile=areaFilter)
    if pointFilter:
        data = data.filter(dome=pointFilter)
    if factoriesFilter:
        data = data.filter(factory_stock=factoriesFilter)
    if productFilter:
        data = data.filter(code_lot=productFilter)

    # Total ritase
    total_ritase = data.count()

    # Total tonase
    total_ton = data.aggregate(total=Sum('tonnage'))['total'] or 0

    # Tonase per material
    lim_total_ton = data.filter(material__iexact='LIM').aggregate(total=Sum('tonnage'))['total'] or 0
    sap_total_ton = data.filter(material__iexact='SAP').aggregate(total=Sum('tonnage'))['total'] or 0

    # Ritase per material
    lim_ritase = data.filter(material__iexact='LIM').count()
    sap_ritase = data.filter(material__iexact='SAP').count()

    return JsonResponse({
        'TotalRitase' : total_ritase,
        'TotalTonnage': total_ton,
        'RitaseLIM'   : lim_ritase,
        'RitaseSAP'   : sap_ritase,
        'TonnageLIM'  : lim_total_ton,
        'TonnageSAP'  : sap_total_ton,
    })

@login_required
@csrf_exempt
def create_quick_selling(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Metode HTTP tidak diizinkan'}, status=405)

    try:
        # Validasi input
        rules = {
            'date_hauling'  : ['required'],
            'shift'         : ['required'],
            'id_material'   : ['required'],
            'id_pile'       : ['required'],
            'unit_code'     : ['required'],
            'code_sub'      : ['required'],
            'code_lot'      : ['required'],
            'barge_code'    : ['required'],
        }

        # Pesan kesalahan validasi
        custom_messages = {
            'date_hauling.required' : 'Date harus diisi.',
            'shift.required'        : 'Shift harus diisi.',
            'id_material.required'  : 'Material harus diisi.',
            'id_pile.required'      : 'Dome harus diisi.',
            'unit_code.required'    : 'Truck harus diisi.',
            'code_sub.required'     : 'SubLot harus diisi.',
            'code_lot.required'     : 'Code Lot harus diisi.',
            'barge_code.required'   : 'Tongkang harus diisi.'
        }

        # Validasi input
        for field, field_rules in rules.items():
            for rule in field_rules:
                if rule == 'required' and not request.POST.get(field):
                    return JsonResponse({'error': custom_messages.get(f'{field}.required', f'{field} wajib diisi')}, status=400)


        # Ambil data dari request
        date_hauling = request.POST.get('date_hauling')
        shift        = request.POST.get('shift')
        id_material  = request.POST.get('id_material')
        id_pile      = request.POST.get('id_pile')
        unit_code    = request.POST.get('unit_code')
        code_sub     = request.POST.get('code_sub')
        code_lot     = request.POST.get('code_lot')
        barge_code   = request.POST.get('barge_code')
        time_hauling = request.POST.get('time_hauling')

        # Ambil nama_material dari tabel Material
        material_obj = Material.objects.filter(id=id_material).first()
        if not material_obj:
            return JsonResponse({'error': 'Material tidak ditemukan'}, status=404)

        nama_material = material_obj.nama_material.upper()

        # Atur type_selling & sale_adjust otomatis
        if nama_material == 'LIM':
            type_selling = 'LIS'
            sale_adjust  = 'HPAL'
        elif nama_material == 'SAP':
            type_selling = 'SAS'
            sale_adjust  = 'SAP'
        else:
            type_selling = None
            sale_adjust  = None

        # Hitung no_urut berdasarkan date_hauling
        last_no = SellingBargingTemp.objects.filter(date_hauling=date_hauling).count()
        no_urut = last_no + 1

        # Hitung code_sub_auto seperti rumus Excel
        count_lot = SellingBargingTemp.objects.filter(code_lot=code_lot).count()
        code_sub_auto = f"SL_{(count_lot // 100 + 1):02d}"

        # Hitung code_inc seperti rumus Excel (filter by code_lot & code_sub_auto)
        count_inc = SellingBargingTemp.objects.filter(
            code_lot=code_lot,
            code_sub_auto=code_sub_auto
        ).count()
        code_inc = (count_inc // 20) + 1

        # Ambil id_stockpile dari relasi dome -> dumping
        dome_obj = SourceMinesDome.objects.filter(id=id_pile).first()
        if not dome_obj:
            return JsonResponse({'error': 'Dome tidak ditemukan'}, status=404)

        id_stockpile = dome_obj.id_dumping  # ambil dari kolom id_dumping

        # Simpan ke database
        with transaction.atomic():
            SellingBargingTemp.objects.create(
                date_hauling  = date_hauling,
                time_hauling  = time_hauling,
                shift         = shift,
                id_material   = id_material,
                id_pile       = id_pile,
                id_stockpile  = id_stockpile,
                unit_code     = unit_code,
                code_sub      = code_sub,
                code_lot      = code_lot,
                barge_code    = barge_code,
                tonnage       = 30.79,
                no_urut       = no_urut,
                type_selling  = type_selling,
                sale_adjust   = sale_adjust,
                code_inc      = code_inc,
                code_sub_auto = code_sub_auto,
                id_user       = request.user.id
            )

        return JsonResponse({'success': True, 'message': 'Data berhasil disimpan.'})

    except Exception as e:
        return JsonResponse({'error': 'Terjadi kesalahan', 'message': str(e)}, status=500)\
        
@login_required
@csrf_exempt
def update_quick_selling(request, id):
    allowed_groups = ['superadmin','admin-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Metode HTTP tidak diizinkan'}, status=405)

    try:
        # Validasi input sama seperti create
        rules = {
            'date_hauling'  : ['required'],
            'shift'         : ['required'],
            'id_material'   : ['required'],
            'id_pile'       : ['required'],
            'unit_code'     : ['required'],
            'code_sub'      : ['required'],
            'code_lot'      : ['required'],
            'barge_code'    : ['required'],
        }
        custom_messages = {
            'date_hauling.required' : 'Date harus diisi.',
            'shift.required'        : 'Shift harus diisi.',
            'id_material.required'  : 'Material harus diisi.',
            'id_pile.required'      : 'Dome harus diisi.',
            'unit_code.required'    : 'Truck harus diisi.',
            'code_sub.required'     : 'SubLot harus diisi.',
            'code_lot.required'     : 'Code Lot harus diisi.',
            'barge_code.required'   : 'Tongkang harus diisi.'
        }

        for field, field_rules in rules.items():
            for rule in field_rules:
                if rule == 'required' and not request.POST.get(field):
                    return JsonResponse({'error': custom_messages.get(f'{field}.required', f'{field} wajib diisi')}, status=400)

        # Ambil object lama
        items = SellingBargingTemp.objects.filter(id=id).first()
        if not items:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

        # Ambil data baru
        date_hauling = request.POST.get('date_hauling')
        shift        = request.POST.get('shift')
        id_material  = request.POST.get('id_material')
        id_pile      = request.POST.get('id_pile')
        unit_code    = request.POST.get('unit_code')
        code_sub     = request.POST.get('code_sub')
        code_lot     = request.POST.get('code_lot')
        barge_code   = request.POST.get('barge_code')
        time_hauling = request.POST.get('time_hauling')

        # Cek material
        material_obj = Material.objects.filter(id=id_material).first()
        if not material_obj:
            return JsonResponse({'error': 'Material tidak ditemukan'}, status=404)

        nama_material = material_obj.nama_material.upper()
        if nama_material == 'LIM':
            type_selling = 'LIS'
            sale_adjust  = 'HPAL'
        elif nama_material == 'SAP':
            type_selling = 'SAS'
            sale_adjust  = 'SAP'
        else:
            type_selling = None
            sale_adjust  = None

        # ✅ FIX: cek id_pile angka atau string kode
        if str(id_pile).isdigit():
            dome_obj = SourceMinesDome.objects.filter(id=id_pile).first()
        else:
            dome_obj = SourceMinesDome.objects.filter(pile_id=id_pile).first()

        if not dome_obj:
            return JsonResponse({'error': 'Dome tidak ditemukan'}, status=404)

        id_pile      = dome_obj.id
        id_stockpile = dome_obj.id_dumping

        # Default pakai nilai lama
        code_sub_auto = items.code_sub_auto
        code_inc      = items.code_inc

        # Kalau code_lot berubah → hitung ulang
        if code_lot != items.code_lot:
            count_lot = SellingBargingTemp.objects.filter(code_lot=code_lot).count()
            code_sub_auto = f"SL_{(count_lot // 100 + 1):02d}"

            count_inc = SellingBargingTemp.objects.filter(
                code_lot=code_lot,
                code_sub_auto=code_sub_auto
            ).count()
            code_inc = (count_inc // 20) + 1

        with transaction.atomic():
            items.date_hauling  = date_hauling
            items.time_hauling  = time_hauling
            items.shift         = shift
            items.id_material   = id_material
            items.id_pile       = id_pile
            items.id_stockpile  = id_stockpile
            items.unit_code     = unit_code
            items.code_sub      = code_sub
            items.code_lot      = code_lot
            items.barge_code    = barge_code
            items.type_selling  = type_selling
            items.sale_adjust   = sale_adjust
            items.code_sub_auto = code_sub_auto
            items.code_inc      = code_inc
            items.id_user       = request.user.id
            items.save()

        return JsonResponse({'success': True, 'message': 'Data berhasil diupdate.'})

    except Exception as e:
        return JsonResponse({'error': 'Terjadi kesalahan', 'message': str(e)}, status=500)


@login_required
def delete_selling_temp(request):
    allowed_groups = ['superadmin','admin-mgoqa','data-control']
    if not request.user.groups.filter(name__in=allowed_groups).exists():
        return JsonResponse(
            {'status': 'error', 'message': 'You do not have permission'}, 
            status=403
    )
    
    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if not job_id:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'}, status=400)
        try:
            job_uuid = UUID(job_id)  # validasi UUID format
            data = SellingBargingTemp.objects.get(id=job_uuid)
            data.delete()
            return JsonResponse({'status': 'deleted'})
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid UUID format'}, status=400)
        except SellingBargingTemp.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Data not found'}, status=404)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@login_required
def get_sale_quick(request, id):
    if request.method == 'GET':
        try:
            items = SellingBargingTemp.objects.get(id=id)
            dumping_point = None
            dome_point    = None
            code_name     = None   

            if items.id_stockpile:
                dumping = SourceMinesDumping.objects.filter(id=items.id_stockpile).first()
                dumping_point = dumping.dumping_point if dumping else None

            if items.id_pile:
                dome = SourceMinesDome.objects.filter(id=items.id_pile).first()
                dome_point = dome.pile_id if dome else None

            product   = SellingCode.objects.filter(product_code=items.code_lot).first()
            code_name = product.product_code if product else None

            data = {
                'id'            : items.id,
                'code_lot'      : items.code_lot, 
                'code_name'     : code_name,
                'barge_code'    : items.barge_code,
                'date_hauling'  : items.date_hauling,
                'time_hauling'  : items.time_hauling,
                'shift'         : items.shift,
                'id_material'   : items.id_material,
                'stockpile'     : dumping_point,
                'dome'          : dome_point,
                'unit_code'     : items.unit_code,
                'tonnage'       : items.tonnage,
                'type_selling'  : items.type_selling,
                'code_inc'      : items.code_inc,
                'code_sub'      : items.code_sub,
                'code_sub_auto' : items.code_sub_auto,
                'sale_adjust'   : items.sale_adjust,
                'no_urut'       : items.no_urut,
                'remarks'       : items.remarks
            }
            return JsonResponse(data)

        except SellingBargingTemp.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required    
@csrf_exempt
def export_sale_data_quick(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    from django.http import HttpResponse

    # Header manual (tetap muncul walaupun data kosong)
    header = [
        'date_barging_in',
        'date_hauling',
        'time',
        'shift',
        'date_barging_load',
        'barge_code',
        'barge_load_loc',
        'barge_unload_loc',
        'no_truck',
        'sale_code',
        'material',
        'stockpile',
        'dome_ori',
        'buyer',
        'code_lot',
        'tonnage',
        'sub_lot',
        'group',
        'adjust_sale',
        'date_barging_out',
        'no_urut',
    ]

    # Mapping header ke field model (kalau tidak ada di model, biarkan None)
    header_field_map = {
        'date_barging_in': None,
        'date_hauling': 'date_hauling',
        'time': 'time_hauling',
        'shift': 'shift',
        'date_barging_load': 'date_hauling',
        'barge_code': 'barge_code',
        'barge_load_loc': None, 
        'barge_unload_loc': None,
        'no_truck': 'unit_code',
        'sale_code': 'type_selling',
        'material': 'material',
        'stockpile': 'stockpile',
        'dome_ori': 'dome',
        'buyer': None,
        'code_lot': 'code_lot',
        'tonnage': 'tonnage',
        'sub_lot': 'code_sub',
        'group': 'code_inc',
        'adjust_sale': 'sale_adjust',
        'date_barging_out': None,
        'no_urut': 'no_urut',
    }

    # Query data (ambil semua dulu supaya bisa isi manual)
    queryset = SellingDetailsBargingTempView.objects.all()

    # Filter tanggal
    from_date = request.GET.get('startDate')
    to_date   = request.GET.get('endDate')
    if from_date and to_date:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        queryset = queryset.filter(date_hauling__range=[from_date, to_date])

    if material := request.GET.get('materialFilter'):
        queryset = queryset.filter(material=material)
    if point := request.GET.get('pointFilter'):
        queryset = queryset.filter(dome=point)
    if product := request.GET.get('productFilter'):
        queryset = queryset.filter(code_lot=product)

    queryset = queryset.annotate(
    shift_order=Case(
        When(shift="Day", then=Value(1)),
        When(shift="Night", then=Value(2)),
        default=Value(3),
        output_field=IntegerField(),
    )
).order_by('-date_hauling', 'shift_order', 'time_hauling', 'no_urut')

    # Buat workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Selling Data Quick'

    # Tulis header
    for col_num, column_title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)

    # Tulis data
    for row_idx, obj in enumerate(queryset, start=2):
        for col_num, column_title in enumerate(header, start=1):
            field_name = header_field_map.get(column_title)
            value = getattr(obj, field_name) if field_name else None
            ws.cell(row=row_idx, column=col_num, value=value)

    # Auto width
    for col_num, col_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_len = len(str(col_title))
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Selling-data-quick.xlsx"'
    wb.save(response)
    return response
