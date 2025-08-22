from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from ...models.selling_barging import SellingBarging
from ...models.selling_details_barging_view import SellingDetailsBargingView
from django.db.models import Sum, Count
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.db.models import Case, When, Value, IntegerField

@login_required
def sale_details_page(request):
    today = datetime.today()
    first_day_of_month = today.replace(day=1)  # Tanggal awal bulan berjalan

    context = {
        'start_date'  : first_day_of_month.strftime('%Y-%m-%d'),
        'end_date'    : today.strftime('%Y-%m-%d'),
    }
    return render(request, 'admin-selling/data/list-selling-details.html',context)

class SellingDetails(View):
    def post(self, request):
        # Ambil semua data invoice yang valid
        data_ore = self._datatables(request)
        return JsonResponse(data_ore, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = SellingDetailsBargingView.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(stockpile__icontains=search) |
                Q(dome__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(code_lot__icontains=search) |
                Q(factory_stock__icontains=search)
            )
       

        # Filter berdasarkan parameter dari request
        from_date       = request.POST.get('startDate')
        to_date         = request.POST.get('endDate')
        materialFilter  = request.POST.get('materialFilter')
        areaFilter      = request.POST.get('areaFilter')
        pointFilter     = request.POST.get('pointFilter')
        factoriesFilter = request.POST.get('factoriesFilter')
        productFilter   = request.POST.get('productFilter')

        if from_date and to_date:
            data = data.filter(date_hauling__range=[from_date, to_date])

        if materialFilter:
            data = data.filter(material=materialFilter)

        if areaFilter:
            data = data.filter(stockpile=areaFilter)

        if pointFilter:
            data = data.filter(dome=pointFilter)

        if factoriesFilter:
            data = data.filter(factory_stock=factoriesFilter)

        if productFilter:
            data = data.filter(code_lot=productFilter)

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"              : item.id,
                "date_barge_in"   : item.date_barge_in,
                "date_hauling"    : item.date_hauling,
                "barge_code"      : item.barge_code,
                "shift"           : item.shift,
                "stockpile"       : item.stockpile,
                "dome"            : item.dome,
                "unit_code"       : item.unit_code,
                "material"        : item.material,
                "code_lot"        : item.code_lot,
                "batch"           : item.batch,
                "code_inc"        : item.code_inc,
                "tonnage"         : item.tonnage,
                "sale_adjust"     : item.sale_adjust,
                "date_barge_out"  : item.date_barge_out,
                "factory_stock"   : item.factory_stock,
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@login_required()
def total_selling(request):
    data = SellingDetailsBargingView.objects.all()

    # Ambil parameter dari request
    from_date       = request.GET.get('startDate')
    to_date         = request.GET.get('endDate')
    materialFilter  = request.GET.get('materialFilter')
    areaFilter      = request.GET.get('areaFilter')
    pointFilter     = request.GET.get('pointFilter')
    factoriesFilter = request.GET.get('factoriesFilter')
    productFilter   = request.GET.get('productFilter')

    # Filter berdasarkan tanggal
    if from_date and to_date:
        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            data = data.filter(date_hauling__range=[from_date, to_date])
        except ValueError:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

    # Filter berdasarkan parameter lain
    if materialFilter:
        data = data.filter(material=materialFilter)
    if areaFilter:
        data = data.filter(stockpile=areaFilter)
    if pointFilter:
        data = data.filter(dome=pointFilter)
    if factoriesFilter:
        data = data.filter(factory_stock=factoriesFilter)
    if productFilter:
        data = data.filter(code_lot=productFilter)

    # Total semua tonase
    total_ton = data.aggregate(total=Sum('tonnage'))['total'] or 0

    # Total berdasarkan type_selling
    lim_total = data.filter(material__iexact='LIM').aggregate(total=Sum('tonnage'))['total'] or 0
    sap_total = data.filter(material__iexact='SAP').aggregate(total=Sum('tonnage'))['total'] or 0

    # Jumlah data
    qty = data.count()

    return JsonResponse({
        'Qty': qty,
        'TotalTonnage': total_ton,
        'TonnageLIM': lim_total,
        'TonnageSAP': sap_total,
    })

@login_required    
@csrf_exempt
def export_sale_data(request):
    # Ambil parameter dari request
    from_date       = request.GET.get('startDate')
    to_date         = request.GET.get('endDate')
    materialFilter  = request.GET.get('materialFilter')
    areaFilter      = request.GET.get('areaFilter')
    pointFilter     = request.GET.get('pointFilter')
    factoriesFilter = request.GET.get('factoriesFilter')
    productFilter   = request.GET.get('productFilter')

    print("Method:", request.method)
    print("GET Params:", request.GET.dict())


    # Buat file Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Selling Data'

    # Header kolom Excel
    header = [
        'No',
        'Date Hauling',
        'Date Barge In',
        'Date Barge Out',
        'Barge Code',
        'Shift',
        'Dome',
        'Stockpile',
        'Material',
        'Truck',
        'Tonnage',
        'Group',
        'Sub Lot',
        'Code Lot',
        'Factory',
        'Type Selling',
        'Sale Adjust',
        'Sale Dome',
    ]

    # Tulis header
    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)

    # Field yang akan diambil dari model
    columns = [
        'date_hauling',
        'date_barge_in',
        'date_barge_out',
        'barge_code',
        'shift',
        'dome',
        'stockpile',
        'material',
        'unit_code',
        'tonnage',
        'code_inc',
        'code_sub',
        'code_lot',
        'factory_stock',
        'type_selling',
        'sale_adjust',
        'sale_dome',
    ]



    queryset = SellingDetailsBargingView.objects.all()

    # Terapkan filter
    if from_date and to_date:
        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            queryset = queryset.filter(date_hauling__range=[from_date, to_date])
        except ValueError:
            return HttpResponse("Format tanggal salah", status=400)

    if materialFilter:
        queryset = queryset.filter(material=materialFilter)
    if areaFilter:
        queryset = queryset.filter(stockpile=areaFilter)
    if pointFilter:
        queryset = queryset.filter(dome=pointFilter)
    if factoriesFilter:
        queryset = queryset.filter(factory_stock=factoriesFilter)
    if productFilter:
        queryset = queryset.filter(code_lot=productFilter)

    # Tambah sorting
    queryset = queryset.annotate(
        shift_order=Case(
            When(shift="Day", then=Value(1)),
            When(shift="Night", then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('-date_hauling', 'shift_order')


    # Baru terakhir values_list
    queryset = queryset.values_list(*columns)

    # Tulis data baris per baris
    for row_num, (row_count, row) in enumerate(enumerate(queryset, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            worksheet.cell(row=row_num + 1, column=col_num).value = cell_value

    # Atur lebar kolom agar otomatis
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Kirim file sebagai response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Selling_data.xlsx"'
    workbook.save(response)
    return response


# @login_required
# def getIdSale(request, id):
#     if request.method == 'GET':
#         try:
#             items = SellingProductions.objects.get(id=id)
#             dumping_point = None
#             dome_point    = None
            
#             if items.id_stockpile:
#                 dumping = SourceMinesDumping.objects.filter(id=items.id_stockpile).first()
#                 if dumping:
#                     dumping_point = dumping.dumping_point

#             if items.id_pile:
#                 dome = SourceMinesDome.objects.filter(id=items.id_pile).first()
#                 if dome:
#                     dome_point = dome.pile_id

#             data = {
#                 'id'                : items.id,
#                 'tgl_hauling'       : items.tgl_hauling, 
#                 'shift'             : items.shift,
#                 'id_material'       : items.id_material,
#                 'id_stockpile'      : items.id_stockpile,
#                 'id_pile'           : items.id_pile,
#                 'dumping_point'     : dumping_point,
#                 'dome_point'        : dome_point,
#                 'id_truck'          : items.unit_code,
#                 'delivery_order'    : items.delivery_order,
#                 'nota'              : items.nota,
#                 'id_factory'        : items.id_factory,
#                 'type_selling'      : items.type_selling,
#                 'empety_weigth_f'   : items.empety_weigth_f,
#                 'fill_weigth_f'     : items.fill_weigth_f,
#                 'netto_weigth_f'    : items.netto_weigth_f,
#                 'id_stock_temp'     : items.id_stock_temp,
#                 'id_dome_temp'      : items.id_dome_temp,
#                 'no_input'          : items.no_input,
#                 'remarks'           : items.remarks,
#                 'batch_g'           : items.batch_g,
#                 'kode_batch_g'      : items.kode_batch_g,
#                 'new_scci'          : items.new_scci,
#                 'new_scci_sub'      : items.new_scci_sub,
#                 'new_kode_batch_scci': items.new_kode_batch_scci,
#                 'new_awk'           : items.new_awk,
#                 'new_awk_sub'       : items.new_awk_sub,
#                 'new_kode_batch_awk': items.new_kode_batch_awk,
#                 'new_batch_awk_pulp': items.new_batch_awk_pulp,
#                 'awk_order'         : items.awk_order,
#                 'scci_order'        : items.scci_order,
#                 'load_code'         : items.load_code,
#                 'haulage_code'      : items.haulage_code,
#                 'date_wb'           : items.date_wb.strftime('%Y-%m-%d %H:%M:%S'),
#                 'timbang_kosong'    : items.timbang_kosong.strftime('%Y-%m-%d %H:%M:%S'),
#                 'timbang_isi'       : items.timbang_isi.strftime('%Y-%m-%d %H:%M:%S'),
#                 'sale_adjust'       : items.sale_adjust,
#                 'sale_dome'         : items.sale_dome,
#             }
#             return JsonResponse(data)
#         except SellingProductions.DoesNotExist:
#             return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

#     return JsonResponse({'error': 'Invalid request method'}, status=400)

# @login_required
# def delete_sale(request):
#     if request.method == 'DELETE':
#         job_id = request.GET.get('id')
#         if job_id:
#             # Lakukan penghapusan berdasarkan ID di sini
#             data = SellingProductions.objects.get(id=int(job_id))
#             data.delete()
#             return JsonResponse({'status': 'deleted'})
#         else:
#             return JsonResponse({'status': 'error', 'message': 'No ID provided'})
#     else:
#         return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    
