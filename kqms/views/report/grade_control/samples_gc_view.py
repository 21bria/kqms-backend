from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.http import HttpResponse
from ....models.sample_gc_view import GradeControlSamples
from django.db.models import Q
from django.shortcuts import render
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views import View
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import locale

# Atur locale untuk menangani format angka dengan koma sebagai pemisah ribuan
locale.setlocale(locale.LC_NUMERIC, 'en_US.UTF-8')

@login_required
def sample_gc_page(request):
    # Dapatkan tanggal hari ini
    today = datetime.today()

    # Hitung tanggal awal minggu (Senin)
    start_of_week = today - timedelta(days=today.weekday())

    # Hitung tanggal akhir minggu (Minggu)
    end_of_week = start_of_week + timedelta(days=6)

    # Menambahkan tanggal awal dan akhir minggu ke konteks
    context = {
        'start_of_week': start_of_week.strftime('%Y-%m-%d'),
        'end_of_week'  : end_of_week.strftime('%Y-%m-%d'),
    }

    return render(request, 'admin-mgoqa/report-gc/list-sample-grade-control.html', context)

class GcSamples(View):
    def post(self, request):
        data_sample = self._datatables(request)
        return JsonResponse(data_sample, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Gunakan fungsi get_joined_data
        data = GradeControlSamples.objects.all()

        if search:
            data = data.filter(
                Q(shift__icontains=search) |
                Q(sampling_area__icontains=search) |
                Q(type_sample__icontains=search) |
                Q(nama_material__icontains=search) |
                Q(sample_method__icontains=search) |
                Q(sample_id__icontains=search) |
                Q(job_number__icontains=search)
            )
       

        # Filter berdasarkan parameter dari request
        from_date      = request.POST.get('from_date')
        to_date        = request.POST.get('to_date')
        materialFilter = request.POST.get('materialFilter')
        method_filter  = request.POST.get('method_filter')
        sourceFilter   = request.POST.get('sourceFilter')


        if from_date and to_date:
            data = data.filter(tgl_sample__range=[from_date, to_date])

        if method_filter:
            data = data.filter(sample_method=method_filter)

        if materialFilter:
            data = data.filter(nama_material=materialFilter)

        if sourceFilter:
            data = data.filter(sampling_area=sourceFilter)


        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Menghitung jumlah total sebelum filter
        records_total = data.count()

        # Menerapkan pagination
        paginator   = Paginator(data, length)
        total_pages = paginator.num_pages

        # Menghitung jumlah total setelah filter
        total_records_filtered = paginator.count

        # Atur paginator
        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        # Format data sesuai dengan DataTables
        data = [
                {
                    "tgl_sample"   : item.tgl_sample,
                    "tgl_deliver"  : item.tgl_deliver,
                    "sampling_area": item.sampling_area,
                    "sample_id"    : item.sample_id,
                    "sample_method": item.sample_method,
                    "nama_material": item.nama_material,
                    "job_number"   : item.job_number,
                    "release_date" : item.release_date,
                    "ni"           : item.ni,
                    "co"           : item.co,
                    "fe2o3"        : item.fe2o3,
                    "fe"           : item.fe,
                    "mgo"          : item.mgo,
                    "sio2"         : item.sio2,
                    "sm"           : item.sm
                } for item in object_list
            ]

        # print(data)

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': total_records_filtered,
            'data'           : data,
            'start'          : start,
            'length'         : length,
            'totalPages'     : total_pages,
        }

@csrf_exempt
def export_gc_sample(request):
    # Ambil filter dari request
    from_date       = request.GET.get('from_date')
    to_date         = request.GET.get('to_date')
    material_filter = request.GET.get('materialFilter')
    method_filter   = request.GET.get('method_filter')

    # Siapkan workbook
    workbook = Workbook()
    ws = workbook.active
    if ws is None:
        raise ValueError("Workbook tidak memiliki worksheet aktif")

    worksheet: Worksheet = ws
    worksheet.title = 'Export Data GC Samples'

    # Header dan kolom
    headers = [
        'No', 'Date sample', 'Deliver', 'Sampling area', 'Sample id',
        'Type sample', 'Sample method', 'Material', 'Job number',
        'Release date', 'Ni', 'Co', 'Fe2O3', 'Fe', 'MgO', 'SiO2'
    ]

    fields = [
        'tgl_sample', 'tgl_deliver', 'sampling_area', 'sample_id',
        'type_sample', 'sample_method', 'nama_material', 'job_number',
        'release_date', 'ni', 'co', 'fe2o3', 'fe', 'mgo', 'sio2'
    ]

    # Tulis header
    for col_num, title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    # Ambil queryset
    queryset = GradeControlSamples.objects.all()
    if from_date and to_date:
        queryset = queryset.filter(tgl_sample__range=[from_date, to_date])
    if method_filter:
        queryset = queryset.filter(sample_method=method_filter)
    if material_filter:
        queryset = queryset.filter(nama_material=material_filter)

    queryset = queryset.values_list(*fields)

    # Tulis data
    for row_index, row in enumerate(queryset, start=2):
        worksheet.cell(row=row_index, column=1, value=row_index - 1)  # No
        for col_num, value in enumerate(row, start=2):
            # Coba konversi nilai numerik dari string jika perlu
            if isinstance(value, str):
                try:
                    value = float(value.replace(',', '').replace(' ', ''))
                except ValueError:
                    pass
            worksheet.cell(row=row_index, column=col_num, value=value)

    # Otomatis lebar kolom
    for col_num, title in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(str(title))
        for cell in worksheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_gc.xlsx"'
    workbook.save(response)
    return response