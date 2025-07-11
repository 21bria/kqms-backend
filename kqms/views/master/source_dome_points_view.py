from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.http import HttpResponse
from ...models.source_model import detailsDome,SourceMinesDome,SourceMinesDumping
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.db import IntegrityError
from django.shortcuts import render
from django.db.models import Q
from django.views.generic import View
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from ...utils.utils import clean_string

@login_required
def sourceDomePoint_page(request):
    return render(request, 'master/list-source-dome-point.html')

class sourceDomePoint_List(View):

    def post(self, request):
        # Ambil semua data 
        minesDome = self._datatables(request)
        return JsonResponse(minesDome, safe=False)

    def _datatables(self, request):
        datatables = request.POST
        # Ambil draw
        draw = int(datatables.get('draw'))
        # Ambil start
        start = int(datatables.get('start'))
        # Ambil length (limit)
        length = int(datatables.get('length'))
        # Ambil data search
        search = datatables.get('search[value]')
        # Ambil order column
        order_column = int(datatables.get('order[0][column]'))
        # Ambil order direction
        order_dir = datatables.get('order[0][dir]')

        # Set record total
        records_total = detailsDome.objects.all().count()
        # Set records filtered
        records_filtered = records_total
        # Ambil semua yang valid
        data = detailsDome.objects.all()

        if search:
            data = detailsDome.objects.filter(
                Q(pile_id__icontains=search) |
                Q(remarks__icontains=search)|
                Q(dumping_point__icontains=search)|
                Q(category__icontains=search)
            )
            records_total = data.count()
            records_filtered = records_total

        # Atur sorting
        if order_dir == 'desc':
            order_by = f'-{data.model._meta.fields[order_column].name}'
        else:
            order_by = f'{data.model._meta.fields[order_column].name}'

        data = data.order_by(order_by)

        # Atur paginator
        paginator = Paginator(data, length)

        try:
            object_list = paginator.page(start // length + 1).object_list
        except PageNotAnInteger:
            object_list = paginator.page(1).object_list
        except EmptyPage:
            object_list = paginator.page(paginator.num_pages).object_list

        data = [
            {
                "id"             : item.id,
                "pile_id"        : item.pile_id,
                "remarks"        : item.remarks,
                "category"       : item.category,
                "status"         : item.status,
                "status_dome"    : item.status_dome,
                "dome_finish"    : item.dome_finish,
                'plan_ni_min'    : item.plan_ni_min,
                'plan_ni_max'    : item.plan_ni_max,
                "dome_finish"    : item.dome_finish,
                "direct_sale"    : item.direct_sale,
                "dumping_point"  : item.dumping_point
            } for item in object_list
        ]

        return {
            'draw'           : draw,
            'recordsTotal'   : records_total,
            'recordsFiltered': records_filtered,
            'data'           : data,
        }

@login_required 
@csrf_exempt
def get_sourceDomePoint(request, id):
    # allowed_groups = ['superadmin','admin-mgoqa','data-control']
    # if not request.user.groups.filter(name__in=allowed_groups).exists():
    #     return JsonResponse(
    #         {'status': 'error', 'message': 'You do not have permission'}, 
    #         status=403
    # )
    if request.method == 'GET':
        try:
            job = SourceMinesDome.objects.get(id=id)

            stockpile = None

            if job.id_dumping:
                dumping = SourceMinesDumping.objects.filter(id=job.id_dumping).first()
                if dumping:
                    stockpile = dumping.dumping_point

            data = {
                'id'          : job.id,
                'pile_id'     : clean_string(job.pile_id), 
                'remarks'     : clean_string(job.remarks),
                'category'    : clean_string(job.category),
                'status'      : clean_string(job.status),
                'plan_ni_min' : job.plan_ni_min,
                'plan_ni_max' : job.plan_ni_max,
                'direct_sale' : clean_string(job.direct_sale),
                'id_dumping'  : job.id_dumping,
                'stockpile'   : stockpile,
                'created_at'  : job.created_at
            }
            return JsonResponse(data)
        except SourceMinesDome.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)

    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def insert_sourceDomePoint(request):
    # allowed_groups = ['superadmin', 'admin-mgoqa','data-control']
    # if not request.user.groups.filter(name__in=allowed_groups).exists():
    #     return JsonResponse(
    #         {'status': 'error', 'message': 'You do not have permission'}, 
    #         status=403
    # )
    if request.method == 'POST':
        pile_id     = request.POST.get('pile_id')
        remarks     = request.POST.get('remarks')
        category    = request.POST.get('category')
        ni_min      = request.POST.get('ni_min')
        ni_max      = request.POST.get('ni_max')
        status      = 1
        direct_sale = request.POST.get('direct_sale')
        id_dumping  = request.POST.get('id_stockpile')

        # print(request.POST)

        # Convert empty strings to None or a default value for numeric fields
        try:
            ni_min = float(ni_min) if ni_min else None
            ni_max = float(ni_max) if ni_max else None
        except ValueError:
            return JsonResponse({'status': 'error', 'message': 'Invalid number format'}, status=400)

        try:
            new_job = SourceMinesDome.objects.create(
                    pile_id     = pile_id,
                    remarks     = remarks,
                    category    = category,
                    plan_ni_min = ni_min,
                    plan_ni_max = ni_max,
                    status      = status,
                    direct_sale = direct_sale,
                    id_dumping  = id_dumping
                    )
            return JsonResponse({
                'status' : 'success',
                'message': 'Data berhasil disimpan.',
                'data': {
                    'id'          : new_job.id,
                    'pile_id'     : new_job.pile_id,
                    'remarks'     : new_job.remarks,
                    'category'    : new_job.category,
                    'status'      : new_job.status,
                    'plan_ni_min' : new_job.plan_ni_min,
                    'direct_sale' : new_job.direct_sale,
                    'id_dumping'  : new_job.id_dumping,
                    'created_at'  : new_job.created_at
                }
            })
        except IntegrityError as e:
            if 'duplicate key' in str(e).lower():
                 return JsonResponse({'status': 'error', 'message': 'Data already exists'}, status=404)
            else:
                return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Metode tidak diizinkan'}, status=405)

@login_required    
def update_sourceDomePoint(request, id):
    # allowed_groups = ['superadmin','admin-mgoqa','data-control']
    # if not request.user.groups.filter(name__in=allowed_groups).exists():
    #     return JsonResponse(
    #         {'status': 'error', 'message': 'You do not have permission'}, 
    #         status=403
    # )
    if request.method == 'POST':
        try:
            # Ambil objek berdasarkan ID
            job = SourceMinesDome.objects.get(id=id)
            
            # Ambil nilai dari POST request dan atur default value jika kosong
            pile_id     = request.POST.get('pile_id')
            remarks     = request.POST.get('remarks')
            category    = request.POST.get('category')
            direct_sale = request.POST.get('direct_sale')
            id_dumping = request.POST.get('stockpile')
            
            # Proses nilai untuk plan_ni_min dan plan_ni_max
            plan_ni_min_str = request.POST.get('ni_min', '')
            plan_ni_max_str = request.POST.get('ni_max', '')
            
            # Convert ke tipe data yang sesuai (misalnya integer) atau atur ke None jika kosong
            try:
                plan_ni_min = float(plan_ni_min_str) if plan_ni_min_str else None
                plan_ni_max = float(plan_ni_max_str) if plan_ni_max_str else None
            except ValueError:
                return JsonResponse({'error': 'Nilai untuk plan_ni_min atau plan_ni_max tidak valid'}, status=400)
            
            # Set nilai ke objek job
            job.pile_id     = pile_id
            job.remarks     = remarks
            job.category    = category
            job.plan_ni_min = plan_ni_min
            job.plan_ni_max = plan_ni_max
            job.direct_sale = direct_sale
            job.id_dumping  = id_dumping
            job.save()

            # Kembalikan respons JSON
            return JsonResponse({
                'id'          : job.id,
                'pile_id'     : job.pile_id,
                'remarks'     : job.remarks,
                'category'    : job.category,
                'plan_ni_min' : job.plan_ni_min,
                'plan_ni_max' : job.plan_ni_max, 
                'direct_sale' : job.direct_sale, 
                'id_dumping'  : job.id_dumping, 
                'created_at'  : job.created_at,
                'updated_at'  : job.updated_at
            })

        except SourceMinesDome.DoesNotExist:
            return JsonResponse({'error': 'Data tidak ditemukan'}, status=404)
        except IntegrityError as e:
            error_message = str(e)
            if 'Duplicate entry' in error_message:
                return JsonResponse({'error': 'Duplikat data: The data already exists', 'message': error_message}, status=400)
            else:
                return JsonResponse({'error': 'Error pada operasi database', 'message': error_message}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Metode tidak diizinkan'}, status=405)

@login_required
def delete_sourceDomePoint(request):
    # allowed_groups = ['superadmin']
    # if not request.user.groups.filter(name__in=allowed_groups).exists():
    #     return JsonResponse(
    #         {'status': 'error', 'message': 'You do not have permission'}, 
    #         status=403
    # )
    if request.method == 'DELETE':
        job_id = request.GET.get('id')
        if job_id:
            data = SourceMinesDome.objects.get(id=int(job_id))
            data.delete()
            return JsonResponse({'status': 'deleted'})
        else:
            return JsonResponse({'status': 'error', 'message': 'No ID provided'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

@csrf_exempt
def export_dome_point(request):
    sourceFilter   = request.POST.get('sourceFilter')

    # workbook = openpyxl.Workbook()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data Dome'

    # Write header row
    header = [
        'No', 
        'Dome', 
        'Stockpile', 
        'Remarks', 
    ]

    for col_num, column_title in enumerate(header, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)  # Mengatur teks menjadi bold

    # List kolom yang ingin diambil
    columns = [
        'pile_id', 
        'dumping_point', 
        'remarks'
    ]

    # Iterator ini mengambil data dalam beberapa bagian, sehingga hemat memori untuk kumpulan data besar.
    queryset = detailsDome.objects.all().values_list(*columns)

    # Membersihkan setiap nilai string dalam hasil queryset
    cleaned_data = [
        [clean_string(value) for value in row]  # Terapkan clean_string ke setiap nilai dalam baris
        for row in queryset
    ]
    
    
    if sourceFilter:
        queryset = queryset.filter(sampling_area=sourceFilter)

    for row_num, (row_count, row) in enumerate(enumerate(cleaned_data, 1), 1):
        worksheet.cell(row=row_num + 1, column=1, value=row_count)
        for col_num, cell_value in enumerate(row, 2):
            cell = worksheet.cell(row=row_num + 1, column=col_num)
            
            if isinstance(cell_value, str):  
                try:
                    cell_value = float(cell_value.replace(',', '').replace(' ', '').replace('.', '', 1))
                except ValueError:
                    pass  # Biarkan tetap string jika bukan angka

            cell.value = cell_value

    # Sesuaikan lebar kolom berdasarkan panjang teks di header
    for col_num, column_title in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        max_length = len(column_title)  # Panjang teks di header
        for row in worksheet.iter_rows(min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dome_point.xlsx"'
    workbook.save(response)

    return response
