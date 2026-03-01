import os
from decimal import Decimal, InvalidOperation
from datetime import datetime, date as date_cls

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from kqms.models import Rainfall, RainfallPoint

# python manage.py import_rainfall data/rainfall.xlsx --dry-run
# python manage.py import_rainfall data/rainfall.xlsx --sheet "Sheet1"

def parse_excel_date(value):
    """
    Terima date dari Excel (datetime/date/str) -> python date.
    Support contoh: 20-Oct-25
    """
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date_cls):
        return value

    if isinstance(value, str):
        v = value.strip()
        # coba beberapa format umum
        for fmt in ("%d-%b-%y", "%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                pass

    return None


def to_decimal(value):
    """
    Convert angka excel/str -> Decimal(2).
    Kosong -> None
    """
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Import rainfall from an .xlsx file. Header row: Date + rainfall point names."

    def add_arguments(self, parser):
        parser.add_argument(
            "filepath",
            nargs="?",
            default="data/rainfall.xlsx",
            help="Path to Excel file. Default: data/rainfall.xlsx",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Sheet name (optional). Default: active sheet",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Only validate & show summary; do not write to DB",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        filepath = options["filepath"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]

        if not os.path.exists(filepath):
            raise CommandError(f"File not found: {filepath}")

        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Ambil header (baris 1)
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if not header or str(header[0]).strip().lower() != "date":
            raise CommandError("Header kolom pertama harus 'Date'.")

        point_names = []
        for h in header[1:]:
            if h is None or str(h).strip() == "":
                point_names.append(None)
            else:
                point_names.append(str(h).strip())

        # Preload / create RainfallPoint untuk setiap header point
        points_by_name = {}
        for name in point_names:
            if not name:
                continue
            obj, _ = RainfallPoint.objects.get_or_create(name=name)
            points_by_name[name] = obj

        created = 0
        updated = 0
        skipped = 0
        invalid_rows = 0

        # iter data mulai baris 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_date = row[0]
            d = parse_excel_date(raw_date)
            if not d:
                # kalau baris kosong total, skip halus
                if all(v in (None, "") for v in row):
                    continue
                invalid_rows += 1
                self.stdout.write(self.style.WARNING(f"Skip row (invalid date): {row}"))
                continue

            # loop tiap kolom point
            for idx, name in enumerate(point_names, start=1):  # start=1 karena row[1] adalah kolom ke-2
                if not name:
                    continue

                mm = to_decimal(row[idx] if idx < len(row) else None)
                if mm is None:
                    skipped += 1
                    continue

                point_obj = points_by_name.get(name)
                if not point_obj:
                    point_obj, _ = RainfallPoint.objects.get_or_create(name=name)
                    points_by_name[name] = point_obj

                # Upsert by (date, point)
                existing = Rainfall.objects.filter(date=d, point=point_obj).first()
                if existing:
                    if existing.milimeter != mm:
                        existing.milimeter = mm
                        if not dry_run:
                            existing.save(update_fields=["milimeter", "updated_at"])
                        updated += 1
                    else:
                        skipped += 1
                else:
                    if not dry_run:
                        Rainfall.objects.create(
                            date=d,
                            point=point_obj,
                            milimeter=mm,
                            description="",
                        )
                    created += 1

        if dry_run:
            # rollback transaction biar DB tidak berubah
            transaction.set_rollback(True)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Import summary"))
        self.stdout.write(f"- created : {created}")
        self.stdout.write(f"- updated : {updated}")
        self.stdout.write(f"- skipped : {skipped}")
        self.stdout.write(f"- invalid : {invalid_rows}")
        self.stdout.write(f"- dry_run : {dry_run}")